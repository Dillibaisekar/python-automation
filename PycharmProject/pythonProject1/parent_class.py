from openpyxl.reader.excel import load_workbook
from selenium.webdriver.common.by import By
from selenium.webdriver import ActionChains

from sel import driver

cust_name=driver.find_element(By.ID,"cus_name")
cust_mobile=driver.find_element(By.ID,"cus_mobile")
cust_mail=driver.find_element(By.ID,"cus_email")
L1_submit=driver.find_element(By.ID,"pf-apply-btn")
otp1 = driver.find_element(By.ID,"otpCode1")
otp2 = driver.find_element(By.ID,"otpCode2")
otp3 = driver.find_element(By.ID, "otpCode3")
otp4 = driver.find_element(By.ID, "otpCode4")
otp5 = driver.find_element(By.ID, "otpCode5")
otp6 = driver.find_element(By.ID, "otpCode6")
resendotp = driver.find_element(By.ID,"pf-otp-btn")
change_mobile = driver.find_element(By.ID,"changeMobNo")
name2 = driver.find_element(By.ID,"cus_name2")
mail2 = driver.find_element(By.ID,"cus_email2")
loan_amount = driver.find_element(By.ID,"cus_loanAmount")
pincode = driver.find_element(By.ID,"cus_pincode2")
form_submit = driver.find_element(By.ID,"pf-apply-btn1")

def getElementById(id):
    elem = driver.find_element(By.ID,id)
    return elem

def picker(value):
    data = driver.find_elements(By.XPATH, "//div[@class='mat-calendar-body-cell-content mat-focus-indicator']")
    for element in data:
        datas = element.text
        print(datas)
        if datas == value:
            element.click()
            break

def excel_data():
    workbook = load_workbook("E:\TW_Screenshots\Test_datas.xlsx")
    s = workbook.active

cust_nameval = excel_data.s.cell(2, 1).value
cust_mobileval = excel_data.s.cell(2, 2).value
cust_mailval = excel_data.s.cell(2, 3).value
name2val = excel_data.s.cell(2, 4).value
mail2val = excel_data.s.cell(2, 5).value
loanamountval = excel_data.s.cell(2, 6).value
pincodeval = excel_data.s.cell(2, 7).value
otp1val = excel_data.s.cell(2,8).value
otp2val = excel_data.s.cell(2, 9).value
otp3val = excel_data.s.cell(2, 10).value
otp4val = excel_data.s.cell(2, 11).value
otp5val = excel_data.s.cell(2, 12).value
otp6val = excel_data.s.cell(2, 13).value
# cust_name.send_keys(nameval)
# cust_mobile.send_keys(mobileval)
# cust_mail.send_keys(mailval)
# otp1.send_keys(otp1val)
# otp2.send_keys(otp2val)
# otp3.send_keys(otp3val)
# otp4.send_keys(otp4val)
# otp5.send_keys(otp5val)
# otp6.send_keys(otp6val)






