import time

import openpyxl
from openpyxl.reader.excel import load_workbook
from selenium import webdriver
from selenium.webdriver import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys

import parent_class


def picker(value):
    data = driver.find_elements(By.XPATH, "//div[@class='mat-calendar-body-cell-content mat-focus-indicator']")
    for element in data:
        datas = element.text
        print(datas)
        if datas == value:
            element.click()
            break

driver = webdriver.Chrome()
driver.get("https://sitsfl.stfc.in/")
driver.maximize_window()
print("sfl")
def main():
    print("main")
    element_to_hover_over = driver.find_element(By.XPATH,
                                            '//*[@id="body"]/app-root/app-header/header/section[2]/div/div/div[2]/p')
    hover = ActionChains(driver).move_to_element(element_to_hover_over)
    hover.perform()
    print("hover done")
    time.sleep(5)
    driver.find_element(By.ID, 'main_nav_cgvf').click()
    time.sleep(3)
    print("Commercial Goods Vehicle Finance loan opened successfully")
    driver.execute_script("window.scrollTo(0, 500)")
    time.sleep(3)

    el = parent_class.getElementById("cus_name")
    el.send_keys("Navin")
    print("name added")

    name1 = driver.find_element(By.ID, "cus_name")
    mobile = driver.find_element(By.ID, "cus_mobile")
    mail = driver.find_element(By.ID, "cus_email")
    workbook = load_workbook("E:\TW_Screenshots\Test_datas.xlsx")
    s = workbook.active
    # nameval = s.cell(2, 1).value
    mobileval = s.cell(2, 2).value
    mailval = s.cell(2, 3).value
    # name1.send_keys(nameval)
    print("name added 2")
    mobile.send_keys(mobileval)
    mail.send_keys(mailval)
    driver.find_element(By.ID, "pf-apply-btn").click()
    print("click")
    time.sleep(5)
    print("OTP screen opened sucessfully")
    driver.find_element(By.ID, "changeMobNo").click()
    mobile.clear()
    mobile.send_keys('9791399723')
    driver.find_element(By.ID, 'pf-apply-btn').click()
    time.sleep(5)
    otp1 = driver.find_element(By.ID, "otpCode1")
    otp1.send_keys('1')
    otp2 = driver.find_element(By.ID, "otpCode2")
    otp2.send_keys('1')
    otp3 = driver.find_element(By.ID, "otpCode3")
    otp3.send_keys('1')
    otp4 = driver.find_element(By.ID, "otpCode4")
    otp4.send_keys('1')
    otp5 = driver.find_element(By.ID, "otpCode5")
    otp5.send_keys('1')
    otp6 = driver.find_element(By.ID, "otpCode6")
    otp6.send_keys('1')
    driver.find_element(By.ID, "otpVerifybtn").click()
    time.sleep(5)
    print("OTP done sucessfully")
    name2 = driver.find_element(By.ID, "cus_name2")
    mail2 = driver.find_element(By.ID, "cus_email2")
    driver.find_element(By.ID, 'loan-dob').click()
    print("Date picker opened")
    driver.find_element(By.XPATH, '//*[@id="mat-datepicker-0"]/mat-calendar-header/div/div/button[1]/span[1]').click()
    print("date picker arrow clicked")
    time.sleep(5)
    picker('1999')
    time.sleep(5)
    picker('MAY')
    picker('21')
    loanamount = driver.find_element(By.ID, "cus_loanAmount")
    pincode = driver.find_element(By.ID, "cus_pincode2")
    name2val = s.cell(2, 4).value
    mail2val = s.cell(2, 5).value
    loanamountval = s.cell(2, 6).value
    pincodeval = s.cell(2, 7).value
    name2.clear()
    name2.send_keys(name2val)
    mail2.clear()
    mail2.send_keys(mail2val)
    loanamount.send_keys(loanamountval)
    pincode.send_keys(pincodeval)
    finalsumbit = driver.find_element(By.ID, "pf-apply-btn1")
    finalsumbit.click()
    print("form submitted successfully")
    driver.close()
main()

def tax():
    driver.get("https://sitsfl.stfc.in/")
    driver.maximize_window()

    element_to_hover_over = driver.find_element(By.XPATH,
                                                '//*[@id="body"]/app-root/app-header/header/section[2]/div/div/div[2]/p')
    hover = ActionChains(driver).move_to_element(element_to_hover_over)
    hover.perform()
    time.sleep(5)
    driver.find_element(By.ID, 'main_nav_cgvf').click()
    time.sleep(3)
    print("Commercial Goods Vehicle Finance loan opened successfully")
    driver.execute_script("window.scrollTo(0, 500)")
    time.sleep(3)
    parent_class.excel_data()
    parent_class.cust_name.send_keys(parent_class.cust_nameval)
    print("name entered successfully")






