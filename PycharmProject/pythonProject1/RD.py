import time

import action as action
from numpy import select
from openpyxl.reader.excel import load_workbook

from selenium import webdriver
from selenium.webdriver import ActionChains, Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select


# fun for find element by ID
def ele(id):
    elem = driver.find_element(By.ID, id)
    return elem
def xpath(xpath):
    elem = driver.find_element(By.XPATH, xpath)
    return elem
def link_text(text):
    elem = driver.find_element(By.LINK_TEXT, text)
    return elem
def text(tex):
    elem = driver.find_element_by_xpath("//span[contains(text(),\+tex)]")
    return elem

# fun for date picker
def picker(value):
        date = driver.find_elements(By.XPATH, "//div[@class='mat-calendar-body-cell-content mat-focus-indicator']")
        for element in date:
            datas = element.text
            # print(datas)
            if datas == value:
                element.click()
                break


def doc_proof_dropdown(value):
    dropdown=Select(driver.find_element(By.ID,'ckyc_failure_depositor_proof_type'))
    dropdown.select_by_value(value)
    return dropdown
def account_type(value):
    dropdown=Select(driver.find_element(By.XPATH,"(//select[@class='form__control ng-untouched ng-pristine ng-invalid'])[1]"))
    dropdown.select_by_value(value)
    return dropdown
def mat_ins(value):
    dropdown = Select(driver.find_element(By.XPATH, "//select[@formcontrolname='inv_maturity_instruction_type']"))
    dropdown.select_by_value(value)
def add_drop_down(value):
    dropdown=Select(driver.find_element(By.XPATH,"//select[@class='form__control ng-untouched ng-pristine ng-invalid']"))
    dropdown.select_by_value(value)



excel=load_workbook("E:\Python Project\Python\PycharmProject\pythonProject1\InputFiles\RD_DATA.xlsx")
data = excel.active
rowcount=data.max_row
columncount=data.max_column
print(rowcount)
print(columncount)
c = webdriver.ChromeOptions()
c.add_argument("--incognito")
driver = webdriver.Chrome(options=c)
driver.get("https://sitsfl.stfc.in/")
driver.maximize_window()
element_hover = driver.find_element(By.XPATH, "//div[@class='head-link'][1]")
ActionChains(driver).move_to_element(element_hover).perform()
time.sleep(3)
ele("main_nav_rd").click()
time.sleep(5)
driver.execute_script("window.scrollTo(0,500)")
time.sleep(3)
mobile = ele('cus_mobile')
pincode = ele('cus_pincode')
amount = ele('cus_investment_amount')
invest_now = ele('pf-apply-btn')

for i in range (2,rowcount+1):
    # for j in range(1, columncount + 1):
    #  print(data.cell(i,1).value)
    mobiledata=(data.cell(i,1).value)
    pincodedata=(data.cell(i,2).value)
    amountdata=(data.cell(i,3).value)
    otp1data=(data.cell(i,4).value)
    otp2data = (data.cell(i, 5).value)
    otp3data = (data.cell(i, 6).value)
    otp4data = (data.cell(i, 7).value)
    otp5data = (data.cell(i, 8).value)
    otp6data = (data.cell(i, 9).value)
    pandata = (data.cell(i,10).value)
    firrstnamedata = (data.cell(i, 11).value)
    lastnamedata = (data.cell(i,12).value)
    maildata = (data.cell(i,13).value)
    investamount2data =(data.cell(i,14).value)
    tenuredata = (data.cell(i,15).value)
    instal_date_data = (data.cell(i,16).value)
    doc_type_data = (data.cell(i,17).value)
    doc_num_data = (data.cell(i,18).value)
    proof_front_data = (data.cell(i,19).value)
    proof_back_data = (data.cell(i, 20).value)
    address1_data = (data.cell(i, 21).value)
    address2_data = (data.cell(i, 22).value)
    depositor_pincode_data = (data.cell(i, 23).value)
    depositor_area_data = (data.cell(i, 24).value)
    upload_photo_data = (data.cell(i,25).value)
    account_number_data = (data.cell(i,26).value)
    confirm_account_number_data = (data.cell(i,27).value)
    ifsc_data = (data.cell(i,28).value)
    acc_name_data = (data.cell(i,29).value)
    cheque_upload_data = (data.cell(i,30).value)
    maturity_ins_data = (data.cell(i,31).value)
    occupation_data = (data.cell(i,32).value)
    father_name_data = (data.cell(i,33).value)
    marital_status_data = (data.cell(i,34).value)
    fd_tenure_data = (data.cell(i,35).value)
    fd_payout_data = (data.cell(i,36).value)
    nom_rel_data = (data.cell(i,36).value)
    nom_title_data = (data.cell(i,36).value)
    nom_fir_name_data = (data.cell(i,36).value)
    nom_last_name_data = (data.cell(i,36).value)


    mobile.send_keys(mobiledata)
    print(mobiledata)
    pincode.send_keys(pincodedata)
    amount.click()
    amount.clear()
    amount.send_keys(amountdata)

    invest_now.click()

    # OTP screen
    otp1 = ele('otpCode1')
    otp2 = ele('otpCode2')
    otp3 = ele('otpCode3')
    otp4 = ele('otpCode4')
    otp5 = ele('otpCode5')
    otp6 = ele('otpCode6')
    otp = ele('otpVerifybtn')
    otp1.send_keys(otp1data)
    otp2.send_keys(otp2data)
    otp3.send_keys(otp3data)
    otp4.send_keys(otp4data)
    otp5.send_keys(otp5data)
    otp6.send_keys(otp6data)
    otp.click()
#     # PAN screen
    pan = xpath('//*[@id="pan2"]')
    pan.send_keys(pandata)
    pan.click()
    time.sleep(10)
    firstname = ele('firstName')
    lastname = ele('lastName')
    dobclick = xpath('//*[@id="depositer-details"]/form/div/div[2]/div/div[3]/div[1]/div[6]/div/input')
    driver.execute_script("window.scrollTo(0, 800)")
    firstname.clear()
    lastname.clear()
    firstname.send_keys(firrstnamedata)
    lastname.send_keys(lastnamedata)
    dobclick.click()
    dobyear = xpath("//button[@class='mat-focus-indicator mat-calendar-period-button mat-button mat-button-base']")
    dobyear.click()
    y=1999
    picker(y)
    time.sleep(10)
    picker('MAY')
    time.sleep(10)
# #     picker('21')
# #     mail = ele('email')
# #     mail.send_keys(maildata)
# #     driver.execute_script("window.scrollTo(0,0)")
# # #
# # # #     xpath for scheme selection
# #     invest_edit = xpath("//i[@class='icon icon-edit'][1]")
# #     invest_edit.click()
# #     investamount2 = ele('name')
# #     investamount2.clear()
# #     time.sleep(1)
# #     print('datacleared sucessfully')
# #     tenure12= xpath("//label[text()='12']")
# #     tenure24= xpath("//label[text()='24']")
# #     tenure36= xpath("//label[text()='36']")
# #     tenure48= xpath("//label[text()='48']")
# #     tenure60= xpath("//label[text()='60']")
# #     tenure36.click()
# #     tenure48.click()
# #     tenure12.click()
# #     tenure24.click()
# #     tenure60.click()
# #     # print('tenure data',tenuredata)
# #     # link_text(tenuredata).click()
# #     installment_date = xpath("(//i[@class='icon icon-edit'])[2]")
# #     installment_date.click()
# #     print(instal_date_data)
# #     picker('27')
# #     print('click')
# #     time.sleep(5)
# #     terms_cond = xpath('//*[@id="depositer-details"]/form/div/div[3]/div[1]/div[4]/div/div[1]/div/div/label')
# #     terms_cond.click()
# #     terms_cond.click()
# #     continue1= xpath("(//button[@class='button button--yellow'])[1]")
# #     continue1.click()
# #     print("1st screen completed sucessfully")
# #     time.sleep(20)
# #     netbanking = xpath("(//a[@class='menu-link'])[2]")
# #     netbanking.click()
# #     time.sleep(5)
# #     bank = xpath("//span[@class='bankLogo lo-nbmahnetbnk']")
# #     bank.click()
# #     time.sleep(5)
# #     make_payment = xpath("(//input[@class='btn'])[2]")
# #     make_payment.click()
# #     time.sleep(20)
# #     driver.execute_script("window.scrollTo(0,500)")
# #     doc_type = ele('ckyc_failure_depositor_proof_type')
# #     doc_type.click()
# #     doc_proof_dropdown('RDAID')
# #     doc_num = xpath("//input[@class='form__control ng-untouched ng-dirty ng-invalid']")
# #     doc_num.send_keys(doc_num_data)
# #     time.sleep(5)
# #     driver.execute_script("window.scrollTo(0,800)")
# #     time.sleep(5)
# #     proof_front = xpath('//*[@id="ckyc_failure_depositor_proof_front"]')
# #     driver.execute_script("arguments[0].click();", proof_front)
# #     time.sleep(5)
# #     # proof_front.click()
# #     proof_front.send_keys(proof_front_data)
# #     proof_back = xpath('//*[@id="ckyc_failure_depositor_proof_back"]')
# #     # proof_back.click()
# #     proof_back.send_keys(proof_back_data)
# #     print('doc uploaded')
# #     continue2 = xpath("//button[@class='button button--yellow button--small']")
# #     continue2.click()
# #     address1 = ele('depositor_address_1')
# #     address1.send_keys(address1_data)
# #
# #     address2 = ele('depositor_address_2')
# #     address2.send_keys(address2_data)
# #     depositor_pincode = ele('depositor_pincode')
# #     depositor_pincode.send_keys(depositor_pincode_data)
# #     ActionChains(driver).key_down(Keys.ARROW_DOWN).pause(2).key_up(Keys.ENTER).perform()
# #     depositor_pincode.send_keys(Keys.ARROW_DOWN)
# #     action.key_down(Keys.COMMAND).key_down("A").key_up(Keys.COMMAND).perform()
# #     action.key_down(Keys.ENTER).perform()
# #     depositor_area = xpath("(//span[@class='mat-option-text'])[1]")
# #     depositor_area.send_keys(Keys.ARROW_DOWN)
# #     depositor_area.send_keys(Keys.ENTER)
# #     depositor_pincode.click()
# #     depositor_pincode.send_keys(Keys.ARROW_DOWN)
# #     depositor_pincode.send_keys(Keys.ENTER)
# #     upload_photo = xpath('//*[@id="depositor_photo"]')
# #     upload_photo.send_keys(upload_photo_data)
# #     continue2.click()
# #     time.sleep(10)
# #     driver.execute_script("window.scrollTo(0,200)")
# #     digital_payment = xpath("//span[@class='switch-handle']")
# #     digital_payment.click()
# #     account_number = xpath("//input[@formcontrolname='bank_accountNumber']")
# #     account_number.clear()
# #     account_number.send_keys(account_number_data)
# #     confirm_account_number = xpath("//input[@formcontrolname='bank_confirmAccountNumber']")
# #     confirm_account_number.send_keys(confirm_account_number_data)
# #     ifsc = xpath("//input[@formcontrolname='bank_IFSC']")
# #     ifsc.send_keys(ifsc_data)
# #     time.sleep(15)
# #     acc_name = xpath("//input[@formcontrolname='bank_personName']")
# #     acc_name.send_keys(acc_name_data)
# #     acc_type = xpath("//select[@formcontrolname='bank_type']")
# #     acc_type.click()
# #     account_type('RDSA')
# #     cheque_upload = xpath("//input[@formcontrolname='bankOCR']")
# #     cheque_upload.send_keys(cheque_upload_data)
# #     continue2 = xpath("//button[@class='button button--yellow button--small']")
# #     continue2.click()
# #
# #      # Additional details screen
# #     maturity_instruction = xpath("//select[@formcontrolname='inv_maturity_instruction_type']")
# #     maturity_instruction.click()
# #     mat_ins("RDCFD")
# #     occupation = xpath("//select[@formcontrolname='inv_occupation_type']")
# #     occupation.click()
# #     add_drop_down("PROFL")
# #     father_name = xpath("//input[@formcontrolname='father_name']")
# #     father_name.send_keys(father_name_data)
# #     marital_status = xpath("//select[@formcontrolname='inv_marital_status']")
# #     marital_status.click()
# #     add_drop_down("RDUNM")
# #     fd_tenure = xpath("//select[@formcontrolname='deposit_fd_tenure_type']")
# #     fd_tenure.click()
# #     add_drop_down("36")
# #     fd_payout = xpath("//select[@formcontrolname='deposit_fd_payout_type']")
# #     fd_payout.click()
# #     add_drop_down("QTR")
# #     print("success")
# # #     Nominee details
# #     nominee = xpath("(//span[@class='switch-handle'])[2]")
# #     nominee.click()
# #     nom_rel = xpath("//select[@formcontrolname='nominee_relationship']")
# #     nom_rel.click()
# #     add_drop_down("FATHR")
# #     # nom_title = xpath("//select[@formcontrolname='nominee_title']")
# #     # nom_title.click()
# #     # add_drop_down("Mr.")
# #     nom_fir_name = xpath("//input[@formcontrolname='nominee_firstName']")
# #     nom_fir_name.send_keys(nom_fir_name_data)
# #     nom_last_name = xpath("//input[@formcontrolname='nominee_lastName']")
# #     nom_last_name.send_keys(nom_last_name_data)
# #     nom_dob = xpath("//input[@formcontrolname='nominee_dob']")
# #     nom_dob.click()
# #     nom_year = xpath("//button[@class='mat-focus-indicator mat-calendar-period-button mat-button mat-button-base']").click()
# #     picker("1994")
# #     time.sleep(5)
# #     picker("OCT")
# #     time.sleep(5)
# #     picker("13")
# #     continue2 = xpath("//button[@class='button button--yellow button--small']")
# #     continue2.click()
# #     print("s")
    break



























